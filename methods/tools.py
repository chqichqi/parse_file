import csv
import json
import logging
import math
from logging.handlers import TimedRotatingFileHandler
import os
import struct
import sys
import time, datetime
import argparse
import openpyxl
from openpyxl import Workbook, load_workbook
# from method.file_manager import FileManager
from tkinter import messagebox
from methods.metaInfo import *
from configparser import ConfigParser
from methods.channels_data import Channels_Data
import numpy as np
# from my_window_item import update_progress
import re

'''
-...JKEHR
-...JKECG（ECG滤波后的文件）
-...JKSTF  (st ecg滤波 2024.04.28新增）
-...JKBRF（呼吸滤波文件 2024.04.29新增）
-...JKBRR（呼吸率文件 2024.04.29新增）
-...JKPPG（PPG波形文件）
-...JKPPR (ppgHR结果文件  2024.04.22新增)
-...JKSTR (st结果文件 2024.04.26新增）
-...JKPPS（房颤 PPG 文件）
-...JKALL（全部sensor数据文件）
-...JKMRK（标记文件）
-...JKGRN（PPG 绿光数据文件）
-...JKAGR（PPG 绿光数据文件 all文件中的ppg）
-...JKBHR（心率数据文件 bhrex）
-...JKAHR（心率数据文件       all文件中的ppg ahrex）
-...JKSTP（计步文件）
-...JKIMU（IMU数据文件
'''
magicCode_dic = {
    'XJKEHR':'XJKEHR-->ECG分析后的R点文件',
    'XJKECG':'XJKECG-->ECG滤波后的文件',
    'XJKSTF':'XJKSTF-->st ecg滤波 2024.04.28新增',
    'XJKBRF':'XJKBRF-->呼吸滤波文件 2024.04.29新增',
    'XJKBRR':'XJKBRR-->呼吸率文件 2024.04.29新增',
    'XJKPPG':'XJKPPG-->PPG波形文件',
    'XJKPPR':'XJKPPR-->ppgHR结果文件  2024.04.22新增',
    'XJKSTR':'XJKSTR-->st结果文件 2024.04.26新增',
    'XJKPPS':'XJKPPS-->房颤PPG文件',
    'XJKALL':'XJKALL-->全部sensor数据文件',
    'XJKMRK':'XJKMRK-->标记文件',
    'XJKGRN':'XJKGRN-->PPG绿光数据文件',
    'XJKAGR':'XJKAGR-->PPG绿光数据文件(all文件中的ppg)',
    'XJKBHR':'XJKBHR-->心率数据文件bhrex',
    'XJKAHR':'XJKAHR-->心率数据文件(all文件中的ppg ahrex)',
    'XJKIMU':'XJKIMU-->IMU数据文件',
    'XJKSTP':'XJKSTP-->计步文件',
}

'''
1---JKwear 1/吉康维尔1
2---JKwear 2/吉康维尔2
3---JKapparel/吉康佩尔
4---贴心集
5---JKcare/吉康恺尔
6---JKcure/吉康护心帖
7---JKsense/吉康心识
'''

device_data_dic = {
    1: '-->JKwear 1/吉康维尔1',
    2: '-->JKwear 2/吉康维尔2',
    3: '-->JKapparel/吉康佩尔',
    4: '-->贴心集',
    5: '-->JKcare/吉康恺尔',
    6: '-->JKcure/吉康护心帖',
    7: '-->JKsense/吉康心识',
}
'''
0-->未加密
1-->服务端加密
2-->设备端加密
'''
encrypt_type_dic = {
    0: '0-->未加密',
    1: '1-->服务端加密',
    2: '2-->设备端加密',
}

'''
0=不连续（开机或者脱落后重新产生的文件）
1=连续，对时成功后分的文件或者手表自动一小时分的文件
2=连续，文件超过1.5小时在读数据时分的文件
3=连续，贴心集两个文件中间手动添加的数据文件
'''
data_coiled_dic = {
    0: '0-->不连续（开机或者脱落后重新产生的文件）',
    1: '1-->连续，对时成功后分的文件或者手表自动一小时分的文件',
    2: '2-->连续，文件超过1.5小时在读数据时分的文件',
    3: '3-->连续，贴心集两个文件中间手动添加的数据文件',
}
'''
0=主机，1= 安卓手机，2= IOS手机，3= PC 客户端 4 = 服务端 5=web端
'''
data_create_dic = {
    0: '0-->主机',
    1: '1-->安卓手机',
    2: '2-->IOS手机',
    3: '3-->PC 客户端',
    4: '4-->服务端',
    5: '5-->web端',
}
'''
0 --- 没有
1--- 有
'''
small_pack_index_dic = {
    0: '0-->没有',
    1: '1-->有',
}

# 最优雅的switch case写法
def switch(case):
    return {
        0: ">",
        1: "<",
        2: "="
    }.get(case,"input err")

def show_message(obj, tip_title, tip_content, tip_type, tip_txt='\n\n请问是否立即打开查看？'):
    # 禁用父窗口
    obj.attributes("-disabled", True)
    # 显示消息框
    if tip_type == 'err':
        result = messagebox.showerror(tip_title, tip_content)
        logging.error(tip_content)
    elif tip_type == 'info':
        result = messagebox.showinfo(tip_title, tip_content)
        logging.info(tip_content)
    elif tip_type == 'ask':
        result = messagebox.askquestion(tip_title, tip_content + tip_txt)
        logging.info(tip_content)
    else:   # 非上述三种情况，就只记录日志，不弹窗提示用户
        logging.info(tip_content)
    # 重新启用父窗口
    obj.attributes("-disabled", False)
    return result

# 获得配置文件数据
def getSysFileData(configFileName):
    # 初始化解析器
    config = ConfigParser()
    try:
        # 读取INI文件
        config.read(configFileName, encoding='utf-8')
        # 获取字典数据
        def get_dictionary(section):
            dict_items = config.items(section)
            return dict(item for item in dict_items)

        # 使用函数获取字典数据
        dic1 = get_dictionary('meta_info_item_dic')
        global meta_info_item_dic
        meta_info_item_dic.update(dic1)
        # print(f'meta_info_item_dic={meta_info_item_dic}')

        dic2 = get_dictionary('small_data_dic')
        stmp_dic = str_to_dic(dic2, configFileName)
        global small_data_dic
        small_data_dic.update(stmp_dic)
        # print(f'small_data_dic={small_data_dic}')

        dic3 = get_dictionary('ext_meta_info_dic')
        stmp_dic = str_to_dic(dic3, configFileName)
        global ext_meta_info_dic
        ext_meta_info_dic.update(stmp_dic)
        # print(f'ext_meta_info_dic={ext_meta_info_dic}')

        dic4 = get_dictionary('setting_item_dic')
        stmp_dic = str_to_dic(dic4, configFileName)
        global setting_item_dic
        setting_item_dic.update(stmp_dic)

        # device_types
        dic5 = get_dictionary('device_types')
        global device_type_dic
        device_type_dic.update(dic5)
        print(f'device_type_dic={device_type_dic}')

        # file_types
        dic6 = get_dictionary('file_types')
        global file_types_dic
        file_types_dic.update(dic6)
        # print(f'file_types_dic={file_types_dic}')

        logging.info(f'读取ini配置文件[{configFileName}]完成...')
        return True
    except Exception as e:
        logging.error(f'读取ini配置文件[{configFileName}]出错啦！原因为：{e}, 将使用[metaInfo.py]文件中的初始值...')
        return False


# 将从ini文件读取的字符串数据，转为字典数据
def str_to_dic(dic_data, configFileName):
    try:
        new_dic = {}
        for k, v in dic_data.items():
            if v == '{}' or v == '':
                new_dic[k] = {}
            else:
                v = v.strip('\n')
                new_dic[k] = eval(v)   # 将字符串形式的字典数据，转为字典类型
        return new_dic
    except Exception as e:
        logging.error(f'ini配置文件[{configFileName}]中的数据存在错误！请检查...')
        return None

def get_file_data(file, type='rb'):
    '''
    获取文件数据
    :param file:
    :param type:
    :return: 返回数组数据
    '''
    try:
        with open(file, type) as f:
            array_data = f.read()
            f.close()
        logging.info(f'读取文件[{file}]数据完成...')
        return array_data
    except Exception as e:
        logging.error(f"打开文件[{file}]异常!原因为：{e}..." )
        return None

def byte_to_int(v):
    '''
    将二进制字节数据，转换为INT数据
    :param v:
    :return:
    '''
    try:
        value = int(v.hex(), 16)
        return value
    except Exception as e:
        logging.error("数据转换异常..." + str(e))
        return None

def format_file_path_list(path_list):
    '''
    格式化文件路径列表数据，使其填充最后的/或\\
    :param path_list:
    :return:
    '''
    new_path_list = []
    if (isinstance(path_list, list)):
        for item in path_list:
            if not (item[-1].__eq__('/') or item[-1].__eq__('\\')):
                new_path_list.append(format_file_path(item))
            else:
                new_path_list.append(item)
    return new_path_list

def format_file_path(path):
    '''
    格式化文件路径数据，使其填充最后的/或\\
    :param path:
    :return:
    '''
    new_path = path
    if not (new_path[-1].__eq__('/') or new_path[-1].__eq__('\\')):
        if str(new_path).__contains__('/'):
            new_path += '/'
        else:
            new_path += '\\'
    return new_path

def get_assign_data(file_data, start_loc, end_loc):
    '''
    获得文件指定数据
    :param file_data:
    :param start_loc: 数据开始位置：数组下标
    :param end_loc: 数据结束位置：数组下标
    :return: 获取的数据
    '''
    try:
        assign_data = byte_to_int(file_data[int(start_loc):int(end_loc)])
        return assign_data
    except Exception as e:
        logging.error("获取指定数据异常..." + str(e))
        return None

def get_file_packLoc(file_data, s=110, e=114):
    '''
    获得文件包位置
    :param file_data:
    :return:list
    '''
    try:
        # 获得文件尾的位置下标[110--113]四个字节，左闭右开
        file_dnd_loc = byte_to_int(file_data[s:e])
        tmp_array = file_data[file_dnd_loc:len(file_data)]
        # 4个字节为一包数据的结束位置
        pack_loc = []
        for i in range(0, len(tmp_array), 4):
            pack_loc.append(byte_to_int(tmp_array[i:(i + 4)]))
        return pack_loc
    except Exception as e:
        logging.error("获取文件尾异常..." + str(e))
        return []

def get_pack_data(file_data, pack_loc, pack_info=None):
    '''
    获得包体数据及包头数据
    :param file_data:
    :param pack_loc:
    :param pack_info:
    :return:list
    '''
    try:
        pack_data_len = byte_to_int(file_data[pack_loc + 16: pack_loc + 20])
        pack_head = file_data[pack_loc: pack_loc + 20]
        pack_data = file_data[pack_loc + 20: pack_loc + 20 + pack_data_len]
        # 当info不为None时，才添加数据，反之则跳过
        if not isinstance(pack_info, type(None)):
            pack_info['pack_head'] = pack_head
            pack_info['pack_data'] = pack_data
            pack_info['pack_data_len'] = pack_data_len
        logging.info("读取包数据长度:" + str(len(pack_data)))
        return pack_data
    except Exception as e:
        logging.error("读包数据异常..." + str(e))
        return []

def get_file_head_data(file_data):
    '''
    获得文件头数据:固定的文件前151个字节的数据
    :return:
    '''
    file_head_data_list = []
    file_head_data_list.append(file_data[0:151])
    return file_head_data_list

def parse_NewProtocol_head(file_data):
    try:
        logging.info('开始解析文件头信息...')
        head_list = []
        head_list.append(f'协议版本：{file_data[0]}')
        magicCode = magicCode_dic[bytesToString(file_data[1:7])]
        head_list.append(f'magicCode：{magicCode}')
        head_list.append(f'commonInfoLen：{hexToInt(bytesToHexString(file_data[7:9]))}')
        device_index = hexToInt(bytesToHexString(file_data[9:10]))
        head_list.append(f'设备类型：{bytesToHexString(file_data[9:10]) + device_data_dic[device_index]}')
        head_list.append(f'用户ID：{bytesToString(file_data[10:42])}')  # 用户ID
        head_list.append(f'设备ID：{bytesToString(file_data[42:55])}')   # 设备编号ID
        file_b_time = hexToInt(bytesToHexString(file_data[55:63]))
        global FILE_BEGIN_TIME
        FILE_BEGIN_TIME = file_b_time
        if file_b_time == 0:
            err_info_list.append(f'文件开始时间为:{file_b_time}，存在异常...')
        datetime_type = datetime.datetime.fromtimestamp(file_b_time // 1000)
        head_list.append(f'文件开始时间：{file_b_time}-->{datetime_type.strftime("%Y-%m-%d %H:%M:%S")}')  # 文件开始时间戳
        file_e_time = hexToInt(bytesToHexString(file_data[63:71]))
        global FILE_END_TIME
        FILE_END_TIME = file_e_time
        if file_e_time == 0:
            err_info_list.append(f'文件结束时间为:{file_e_time}，存在异常...')
        elif (file_e_time - file_b_time) < 0:
            err_info_list.append(f'文件结束时间小于文件开始时间，存在异常...')
        datetime_type = datetime.datetime.fromtimestamp(file_e_time // 1000)
        head_list.append(f'文件结束时间：{file_e_time}-->{datetime_type.strftime("%Y-%m-%d %H:%M:%S")}')  # 文件结束时间戳
        head_list.append(f'sensor软件版本：{bytesToHexString(file_data[71:73])}')
        head_list.append(f'sensor硬件版本：{bytesToHexString(file_data[73:74])}')
        encrypt_type = hexToInt(bytesToHexString(file_data[74:75]))
        head_list.append(f'加密类型：{encrypt_type_dic[encrypt_type]}')
        head_list.append(f'数据是否连续：{data_coiled_dic[hexToInt(bytesToHexString(file_data[75:76]))]}')
        head_list.append(f'数据创建端：{data_create_dic[hexToInt(bytesToHexString(file_data[76:77]))]}')
        head_list.append(f'小包数据是否有序号：{small_pack_index_dic[hexToInt(bytesToHexString(file_data[77:78]))]}')
        logging.info('解析文件头信息结束...')
        return head_list, encrypt_type, magicCode
    except Exception as e:
        logging.error(f'解析文件头信息出错啦！原因为：{e}...')
        return None, None

def parse_MetaInfo(file_data, magicCode, s_location=0):
    try:
        logging.info(f'开始解析文件头MetaInfo信息...')
        metaInfo_item_list = []
        metaInfo_item_dic = {}
        metaExtInfo_dic = {}
        metaInfo_list = []
        channels_list = []
        metaExtInfo_list = []
        node_data = {}
        metaInfo_data_dic = {}
        metaInfo_length = hexToInt(bytesToHexString(file_data[s_location:s_location+2]))
        s_location += 2
        metaInfo_length -= 2
        while metaInfo_length > 0:
            metaInfo_type = ''
            metaInfo_version = 0
            index = 0
            # 解析基础的metaInfo
            for k, v in base_meta_info_dic.items():
                value = hexToInt(bytesToHexString(file_data[s_location:s_location+int(v)]))
                # metaInfo_item_list.append(f'[{k}：{value}]')
                metaInfo_data_dic[f'data{index+1}'] = f'{k}：{value}'
                if index == 0:
                    metaInfo_type = str(value)
                    metaInfo_item_dic['type'] = str(value)
                    metaExtInfo_dic['type'] = str(value)
                elif index == 1:
                    metaInfo_version = value
                    metaInfo_item_dic['version'] = value
                elif index == 3:
                    metaInfo_item_dic['offset'] = value
                elif index == 4:
                    metaInfo_item_dic['dataLen'] = value
                s_location += int(v)
                index += 1
                metaInfo_length -= int(v)
            # 将metaInfo数据缓存起来，供解析包数据时使用
            metaInfo_item_list.append(metaInfo_item_dic.copy())
            metaInfo_item_dic.clear()
            # 解析扩展的metaInfo
            if ext_meta_info_dic.get(metaInfo_type) is None:
                err_info_list.append(f'metaInfo_type:{metaInfo_type}不存在，请先根据《文件协议》之规定添加后重试...')
            stmp_dic = ext_meta_info_dic[metaInfo_type]
            if stmp_dic.get(metaInfo_version) is None:
                if len(stmp_dic) > 0:   # 存在一种情况：当前metaInfo_type没有扩展的版本，故就为空
                    err_info_list.append(f'metaInfo_type:{metaInfo_type}中版本:{metaInfo_version}缺少扩展配置信息，'
                                     f'请先根据《文件协议》之规定添加后重试...')
            else:
                tmp_dic = stmp_dic[metaInfo_version].copy()
                for k, v in tmp_dic.items():
                    if '采样率' in k or 'Ffs' in k or 'ffs' in k:   # 这里需要处理采样率的转换:于2024-9-20 10点已处理
                        sample = bytesToFloat(file_data[s_location:s_location+int(v)])
                        # 只有是ppg相关的文件才验证这个采样率
                        if (metaInfo_type == '15' or metaInfo_type == '8' or metaInfo_type == '9') and \
                                not('agr' in str(magicCode).lower() or 'ahr' in str(magicCode).lower()):
                            if (float(sample[0]) < (PPG_FFS - float(setting_item_dic['ppg_ffs'])) or
                                    float(sample[0]) > (PPG_FFS + float(setting_item_dic['ppg_ffs']))):
                                err_info_list.append(f'PPG采样率为：{sample[0]}，'
                                                     f'超出了统计设置[{PPG_FFS}±{setting_item_dic["ppg_ffs"]}]的误差范围...')
                        metaInfo_data_dic[f'data{index+1}'] = f'{k}：{sample[0]}'
                        metaExtInfo_dic[k] = sample[0]
                    elif k == 'text':   # 版本信息，需要处理掉\n
                        version_data = str(bytesToString(file_data[s_location:s_location+int(v)])).replace('\n',',')
                        metaInfo_data_dic[f'data{index+1}'] = f'{k}：{version_data}'
                        metaExtInfo_dic[k] = version_data
                    elif '导联顺序' in k or '导联序号' in k or 'channel' in str(k).lower():
                        d = file_data[s_location:s_location+int(v)]
                        # 缓存导联数据
                        for i in range(len(d)):
                            if d[i] != 0:
                                channels_list.append(Channels_Data.channel_dic[d[i]])
                        metaInfo_data_dic[f'data{index+1}'] = f'{k}：{channels_list.copy()}'
                        metaExtInfo_dic[k] = channels_list.copy()
                        channels_list.clear()
                    else:
                        metaInfo_data_dic[
                            f'data{index+1}'] = f'{k}：{hexToInt(bytesToHexString(file_data[s_location:s_location+int(v)]))}'
                        metaExtInfo_dic[k] = hexToInt(bytesToHexString(file_data[s_location:s_location+int(v)]))
                    s_location += int(v)
                    index += 1
                    metaInfo_length -= int(v)
            if 'V' in meta_info_item_dic[metaInfo_type]:
                node_data['node'] = meta_info_item_dic[metaInfo_type]+str(metaInfo_version)
            else:
                node_data['node'] = meta_info_item_dic[metaInfo_type]
            metaExtInfo_list.append(metaExtInfo_dic.copy())
            metaExtInfo_dic.clear()
            node_data.update(metaInfo_data_dic.copy())
            metaInfo_data_dic.clear()
            metaInfo_list.append(node_data.copy())
            node_data.clear()
        logging.info(f'解析文件头MetaInfo信息结束...')
        return metaInfo_list, metaInfo_item_list, metaExtInfo_list
    except Exception as e:
        err = f'解析文件头MetaInfo信息出错啦！原因为：metaType={e}不存在...'
        logging.error(err)
        return None, None, None

# 解析包数据
def parse_PackData(data, data_len, item_data_list, metaExtInfo_list=None):
    try:
        bigPack_head_list = []
        node_data_dic = {}
        packHead_data_dic = {}
        index = 0
        step_flg = False
        step_sum_data = 0      # 计步数据和
        last_step_time = 0
        current_step_data = 0
        ehr_flg = False
        ehr_sum_flg = False    # EHR房颤统计标记
        ehr_af_slice_sum_list = []  # EHR房颤片段统计
        bhr_flg = False
        bhr_heart_fast_count = 0     # bhr心动过速总次
        bhr_heart_slow_count = 0     # bhr心动过缓总次
        bhr_heart_max = 0            # bhr最大心率
        bhr_heart_min = 200          # bhr最小心率
        bhr_isValidate_time = 0
        bhr_af75_sum_time = 0
        bhr_af50_sum_time = 0        # 实际统计
        bhr_af_sum_time = 0
        bhr_af_flg = False
        bhr_af_slice_sum_list = []        # BHR房颤片段统计
        bigPack_e_time = 0
        bigPack_idx = 1     # 大包序号
        logging.info(f'开始解析文件包体数据...')
        while data_len > 0:
            # 解析包头信息:20bytes
            bigPack_b_time = hexToInt(bytesToHexString(data[index:index+8]))
            if bigPack_idx == 1:
                if bigPack_b_time != FILE_BEGIN_TIME:
                    err_info_list.append(f'大包{bigPack_idx}的开始时间:{bigPack_b_time}，不等文件开始时间:{FILE_BEGIN_TIME}...')
            else:
                if (bigPack_b_time - bigPack_e_time) != 0:
                    err_info_list.append(f'大包{bigPack_idx}的开始时间:{bigPack_b_time}，不等前一包的结束时间:{bigPack_e_time}...')
            datetime_type = datetime.datetime.fromtimestamp(bigPack_b_time // 1000)
            packHead_data_dic['data1'] = f'大包开始时间：{bigPack_b_time}-->{datetime_type.strftime("%Y-%m-%d %H:%M:%S")}'
            bigPack_e_time = bigPack_b_time + hexToInt(bytesToHexString(data[index+8:index+12]))
            if (bigPack_b_time - bigPack_e_time) == 0:
                err_info_list.append(f'大包{bigPack_idx}的开始时间:{bigPack_b_time}，等于结束时间:{bigPack_e_time}...')
            datetime_type = datetime.datetime.fromtimestamp(bigPack_e_time // 1000)
            packHead_data_dic['data2'] = f'大包数据时长：{hexToInt(bytesToHexString(data[index+8:index+12]))}ms' \
                                         f'-->包结束时间：{datetime_type.strftime("%Y-%m-%d %H:%M:%S")}'
            packData_len = hexToInt(bytesToHexString(data[index+12:index+16]))
            packHead_data_dic['data3'] = f'原始数据长度：{packData_len}'
            encrypt_packData_len = hexToInt(bytesToHexString(data[index+16:index+20]))
            packHead_data_dic['data4'] = f'加密数据长度：{encrypt_packData_len}'
            # 如果未加密数据长度为0，则使用加密数据长度
            if packData_len == 0:
                err_info_list.append(f'大包{bigPack_idx}的原始数据长度为:0，将使用加密数据长度:{encrypt_packData_len}进行解析...')
                packData_len = encrypt_packData_len
            index += 20
            data_len -= 20
            # 解析包体数据
            table_pack_head = ''
            for item in item_data_list:
                if item['dataLen'] > 0:
                    small_table_head = small_data_dic[meta_info_item_dic[item["type"]]]
                    if len(small_table_head) > 0:
                        table_pack_head += meta_info_item_dic[item["type"]] + '=>{'
                        for k in small_table_head[item["version"]].keys():
                            if '&' in k:
                                s_v = k.split('&')
                                table_pack_head += (s_v[0] + ' '+ s_v[1] + ' ')
                            elif '与' in k:
                                s_v = k.split('与')
                                table_pack_head += (s_v[0] + ' ' + s_v[1] + ' ')
                            elif 'and' in k:
                                s_v = k.split('and')
                                table_pack_head += (s_v[0] + ' ' + s_v[1] + ' ')
                            else:
                                table_pack_head += (k +' ')
                        table_pack_head = table_pack_head[0:len(table_pack_head)-1] + '} '
                    else:
                        # 处理一下多通道数据，使其按每个通道进行归整一下
                        if not (metaExtInfo_list is None):
                            for v_dic in metaExtInfo_list:
                                if item["type"] == v_dic['type']:
                                    if '导联顺序' in v_dic:
                                        key = '导联顺序'
                                    elif '导联序号' in v_dic:
                                        key = '导联顺序'
                                    if ('导联顺序' in v_dic or '导联序号' in v_dic) and len(v_dic[key]) > 0:
                                        stmp = ''
                                        for v in v_dic[key]:
                                            stmp += (v + ' ')
                                        table_pack_head += meta_info_item_dic[item["type"]] + '=>{' + stmp[0:len(stmp)-1] + '} '
                                    else:
                                        table_pack_head += meta_info_item_dic[item["type"]] + ' '
                                    break
                        else:
                            table_pack_head += meta_info_item_dic[item["type"]]+' '
            packHead_data_dic['data5'] = f'小包数据：{table_pack_head}'
            ismall = 6
            packData_len2 = packData_len
            logging.info(f'开始解析大包{bigPack_idx}中的小包数据...')
            ehr_af_slice_time = 0
            ehr_af_slice_begin_time = 0
            ehr_af_slice_sum_dic = {}  # EHR房颤片段统计
            bhr_af_slice_time = 0
            bhr_af_slice_begin_time = 0
            bhr_af_slice_sum_dic = {}  # EHR房颤片段统计
            while packData_len > 0:
                small_data_list = []
                for item in item_data_list:
                    if item['dataLen'] > 0:
                        value = data[index:index+item['dataLen']]
                        stmp_list = parse_SmallPackData(item['type'], item['version'], value)
                        small_data_list.append(stmp_list.copy())
                        index += item['dataLen']
                        packData_len -= item['dataLen']
                        # 计步文件
                        if item['type'] == '7':  # 计步文件中的步数
                            step_sum_data += stmp_list[0]
                            current_step_data = stmp_list[0]   # 缓存当前步数
                            step_flg = True
                        if item['type'] == '13':  # 计步文件中的更新步数所对应的时间
                            new_list = str(stmp_list[0]).split('-->')
                            if int(new_list[0]) < FILE_BEGIN_TIME:
                                err_info_list.append(f'大包{bigPack_idx}中存在步数：{current_step_data}的发生时间：{new_list[0]}'
                                                     f'早于文件开始时间：{FILE_BEGIN_TIME}了！...')
                            if int(new_list[0]) > FILE_END_TIME:
                                err_info_list.append(f'大包{bigPack_idx}中存在步数：{current_step_data}的发生时间：{new_list[0]}'
                                                     f'晚于文件结束时间：{FILE_END_TIME}了！...')
                            if last_step_time != 0 and last_step_time > int(new_list[0]):
                                err_info_list.append(f'大包{bigPack_idx}中存在步数：{current_step_data}的发生时间：{new_list[0]}'
                                                     f'小于上个步数的发生时间：{last_step_time}了！...')
                            last_step_time = int(new_list[0])

                        # bhr数据
                        if item['type'] == '8':
                            hrInterval = 60000 // int(stmp_list[0])
                            if hrInterval >= int(setting_item_dic['heart_fast']):
                                bhr_heart_fast_count += 1
                            elif hrInterval <= int(setting_item_dic['heart_slow']):
                                bhr_heart_slow_count += 1
                            if bhr_heart_max < hrInterval:
                                bhr_heart_max = hrInterval
                            if bhr_heart_min > hrInterval:
                                bhr_heart_min = hrInterval
                            # 房颤置信度，需要version：1才能统计
                            # stmp_list中元素0：间期，1：波形质量，2：房颤标记，3：全局序号，4：房颤置信度
                            if int(item['version']) >= 1 and len(stmp_list) > 4:
                                if int(stmp_list[4]) >= int(setting_item_dic['ppg_conf']) and not bhr_af_flg:
                                    bhr_af_flg = True
                                if bhr_af_flg and int(stmp_list[1]) == 0 and int(stmp_list[2]) == 2:
                                    bhr_af75_sum_time += int(stmp_list[0])  # 间期就是时间
                                if int(stmp_list[1]) == 0:  # PPG波形质量，0=好1=差
                                    bhr_isValidate_time += int(stmp_list[0])
                                if int(stmp_list[4]) >= 50 and int(stmp_list[1]) == 0:
                                    bhr_af50_sum_time += int(stmp_list[0])  # 间期就是时间
                            else:
                                if int(stmp_list[1]) == 0 and int(stmp_list[2]) == 2:
                                    bhr_af_sum_time += int(stmp_list[0])  # 间期就是时间
                                    bhr_af_flg = True
                                if int(stmp_list[1]) == 0:  # PPG波形质量，0=好1=差
                                    bhr_isValidate_time += int(stmp_list[0])
                            # bhr房颤片段统计
                            if int(stmp_list[2]) == 2:
                                if bhr_af_slice_begin_time == 0:
                                    bhr_af_slice_begin_time = int(stmp_list[3])
                                bhr_af_slice_time += int(stmp_list[0])
                            if int(stmp_list[2]) != 2:
                                if int(bhr_af_slice_time / 1000) >= int(setting_item_dic['ppg_af_slice']):
                                    bhr_af_slice_sum_dic['begin'] = bhr_af_slice_begin_time
                                    bhr_af_slice_sum_dic['len'] = (int(bhr_af_slice_time / 1000))
                                    bhr_af_slice_sum_list.append(bhr_af_slice_sum_dic.copy())
                                    bhr_af_slice_sum_dic.clear()
                                bhr_af_slice_time = 0
                                bhr_af_slice_begin_time = 0
                            bhr_flg = True
                        # ehr数据
                        if item['type'] == '12' or item['type'] == '14':
                            if '(AFIB' in stmp_list[3] and \
                                    stmp_list[5] > int(setting_item_dic['ecg_af']) and (not ehr_sum_flg):
                                sum_info_list.append(f'ehr数据按设置的af置信度≥{setting_item_dic["ecg_af"]}%'
                                                     f'进行统计为:{stmp_list[3]}...')
                                ehr_sum_flg = True
                            ehr_flg = True
                packHead_data_dic[f'data{ismall}'] = small_data_list.copy()
                ismall += 1
            # 处理最后一个连续片段
            if int(bhr_af_slice_time / 1000) >= int(setting_item_dic['ppg_af_slice']):
                bhr_af_slice_sum_dic['begin'] = bhr_af_slice_begin_time
                bhr_af_slice_sum_dic['len'] = (int(bhr_af_slice_time / 1000))
                bhr_af_slice_sum_list.append(bhr_af_slice_sum_dic.copy())
            if int(ehr_af_slice_time / 1000) >= int(setting_item_dic['ecg_af_slice']):
                ehr_af_slice_sum_dic['begin'] = ehr_af_slice_begin_time
                ehr_af_slice_sum_dic['len'] = (int(ehr_af_slice_time / 1000))
                ehr_af_slice_sum_list.append(ehr_af_slice_sum_dic.copy())
            logging.info(f'解析大包{bigPack_idx}中的小包数据结束...')
            data_len -= packData_len2
            node_data_dic['node'] = f'大包{bigPack_idx}'
            node_data_dic.update(packHead_data_dic.copy())
            packHead_data_dic.clear()
            bigPack_head_list.append(node_data_dic.copy())
            node_data_dic.clear()
            bigPack_idx += 1
        # 处理计步总和数据
        if step_flg:
            if step_sum_data > 0:
                sum_info_list.append(f'步数总和为:{step_sum_data}步')
        # 最后还得判断一下最后一包的结束时间是否与文件结束时间一致
        if (FILE_END_TIME - bigPack_e_time) != 0:
            err_info_list.append(f'大包{bigPack_idx-1}的结束时间:{bigPack_e_time}，不等文件的结束时间:{FILE_END_TIME}...')
        # 处理bhr相关数据
        if bhr_flg:
            if bhr_af_flg:
                sum_info_list.append(f'bhr数据分析为：疑似房颤...')
            else:
                sum_info_list.append(f'bhr数据分析为：未见房颤...')
            # af片段统计
            af_slice_sum(bhr_af_slice_sum_list, sum_info_list, 'bhr', 'ppg_af_slice')
            if bhr_af75_sum_time > 0:  # 表示数据中有置信度
                vv = float(bhr_af75_sum_time / bhr_isValidate_time)
                if (vv*100) >= int(setting_item_dic['ppg_eff_per']):
                    exc_pre = round(float(bhr_af50_sum_time / bhr_isValidate_time), 5) * 100
                    comp_v = '大于等于'
                else:
                    comp_v = '小于'
                    exc_pre = '0'
                sum_info_list.append(f'bhr数据中房颤置信度≥75%的有效占比为：{round(vv*100, 5)}%'
                                         f'{comp_v}设置的有效占比：{setting_item_dic["ppg_eff_per"]}%，'
                                         f'故按房颤置信度≥50%的真实负荷进行统计占比为：{exc_pre}%...')
            else:
                exc_pre = round(float(bhr_af_sum_time / bhr_isValidate_time),5) * 100
                sum_info_list.append(f'bhr中为旧版数据，没有房颤置信度数据，故异常节律占比为：{exc_pre}%...')
            sum_info_list.append(f'bhr数据中最大脉率为：{bhr_heart_max}次/分...')
            sum_info_list.append(f'bhr数据中最小脉率为：{bhr_heart_min}次/分...')
            sum_info_list.append(f'bhr数据中脉率≥{setting_item_dic["heart_fast"]}次/分，共发生了:{bhr_heart_fast_count}次...')
            sum_info_list.append(f'bhr数据中脉率≤{setting_item_dic["heart_slow"]}次/分，共发生了:{bhr_heart_slow_count}次...')
        if ehr_flg:
            if not ehr_sum_flg:
                sum_info_list.append(f'ehr数据按设置的af置信度≥{setting_item_dic["ecg_af"]}%进行统计为：未见房颤...')
            # af片段统计
            # af_slice_sum(ehr_af_slice_sum_list, sum_info_list, 'ehr', 'ecg_af_slice')
        logging.info(f'解析文件包体数据结束...')
        return bigPack_head_list
    except Exception as e:
        err = f'解析文件包体数据出错啦！原因为：{e}...'
        logging.error(err)
        # err_info_list.append(err)
        return None

def af_slice_sum(obj_list, sum_obj, tip_txt, slice_name):
    sum_obj.append(f'{tip_txt}数据中发生持续时长≥{setting_item_dic[slice_name]}秒房颤片段{len(obj_list)}次...')
    if tip_txt == 'bhr':
        s = '全局序号'
    else:
        s = 'R点'
    if len(obj_list) > 0:
        i = 1
        for item in obj_list:
            s1 = ''
            s2 = ''
            for k, v in item.items():
                if k == 'begin':
                    s1 = f'从{s}位置：{v}开始,'
                elif k == 'len':
                    s2 = f'持续时长：{v}秒'
            sum_obj.append(f'房颤片段{i}-->{s1}{s2}...')
            i += 1

# 解析包体中小包数据内容
def parse_SmallPackData(data_type, version, data):
    data_type_name = meta_info_item_dic[data_type]
    data_type_dic = small_data_dic[data_type_name]
    s_tmp = []
    idx = 0
    try:
        if isinstance(data_type_dic, dict):
            if 0 in data_type_dic.keys():  # 表示包含版本0或1
                data_dic = data_type_dic[version].copy()
            else:
                data_dic = data_type_dic.copy()
        if len(data_dic) > 0:
            for k, v in data_dic.items():
                if '&' in k or '与' in k or 'and' in k:  # 波形质量&房颤标识,还需要单独处理
                    s = bytesToHexString(data[idx:idx + v])
                    s_tmp.append(s[0])
                    s_tmp.append(s[1])
                else:
                    if k == '心拍相关疾病' or str(k).__contains__('心拍'):
                        stmp = (data[idx:idx + v]).decode('ascii')
                    elif k == '节律相关疾病' or str(k).__contains__('节律'):
                        stmp = RhythmDisType[hexToInt(bytesToHexString(data[idx:idx + v]))]
                    else:
                        stmp = hexToInt(bytesToHexString(data[idx:idx + v]))
                    if 'channel' in k or '通道数' in k:
                        stmp = Channels_Data.channel_dic[stmp]   # 将数据直接转换为通道数
                    s_tmp.append(stmp)
                idx += v
        else:
            if data_type_name == 'ppg-hr结果文件' or data_type_name == 'grn' \
                    or data_type_name == '呼吸阻抗' or str(data_type_name).__contains__('ecg滤波后数据'):
                s_tmp = bytesToFloat(data)  # bytesToFloat函数返回就是list对象
            elif data_type_name == '时间戳':  # 如果是时间戳，则将其转换为：YYYYMMDD hh:mm:ss
                bigPack_b_time = hexToInt(bytesToHexString(data))
                v = int(timestampToDate(bigPack_b_time))
                s_tmp.append(f'{bigPack_b_time}-->{v}')
            # ecg按2字节整型数据进行转换
            elif 'ecg' in data_type_name:
                s_tmp = bytes_To_TowInt(data)
            else:  # 计步数据在此处理
                v = hexToInt(bytesToHexString(data))
                s_tmp.append(v)
    except Exception as e:
        err = f'解析文件包体数据小包数据出错啦！原因为：{e}...'
        logging.error(err)
        err_info_list.append(err)
    return s_tmp

def validate_fileIsEncrypt(file_data):
    if int(file_data[0]) < 10:  # 旧协议
        ty = 'old'
    else:
        ty = 'new'
    encr_loc_s = int(file_Protocol_dic[ty]['enc_loc'][0])
    encr_loc_e = encr_loc_s + int(file_Protocol_dic[ty]['enc_loc'][1])
    encrypt_flg = hexToInt(bytesToHexString(file_data[encr_loc_s:encr_loc_e]))
    if encrypt_flg != 0:  # 已加密
        return True, ty
    else:
        return False, ty

def update_file_data(file_data, b_dataArray, loc_s, loc_e):
    # file_data[user_id_loc_s: user_id_loc_e] = bytes(b_data)
    new_data = file_data[0:loc_s]
    new_data += b_dataArray
    new_data += file_data[loc_e:]
    return new_data

def get_data_loc(update_item, p_ty='new'):
    loc_s = int(file_Protocol_dic[p_ty][update_item][0])
    loc_e = loc_s + int(file_Protocol_dic[p_ty][update_item][1])
    return loc_s, loc_e

def modify_file_time(file_data, modify_data, p_ty, modify_ty, file_type):
    try:
        new_file_data = file_data
        # 文件开始时间
        logging.info('正在修改文件[开始时间]...')
        loc_s, loc_e =  get_data_loc('file_b_time_loc', p_ty)
        begin_time = file_data[loc_s:loc_e]
        old_begin_timestamp = int.from_bytes(begin_time, byteorder='big')
        b_beginTime_array, new_begin_timestamp = get_time_data(file_data[loc_s:loc_e], modify_data, modify_ty)
        new_file_data = update_file_data(new_file_data, b_beginTime_array, loc_s, loc_e)
        # 文件结束时间
        logging.info('正在修改文件[结束时间]...')
        loc_s, loc_e = get_data_loc('file_e_time_loc', p_ty)
        end_time = file_data[loc_s:loc_e]
        old_end_timestamp = int.from_bytes(end_time, byteorder='big')
        # 新的结束时间为：新的开始时间戳+（旧的结束时间戳-旧的开始时间戳）
        stmp = old_end_timestamp - old_begin_timestamp
        # 将结束时间转换为日期+时间
        file_end_time = timestampToDate(new_begin_timestamp + stmp)
        print(f'file_end_time={file_end_time}')
        b_endTime_array = inttimestampTobytes(new_begin_timestamp + stmp)
        new_file_data = update_file_data(new_file_data, b_endTime_array, loc_s, loc_e)
        # 处理每包中的开始时间
        pack_data_time_loc_list = get_packData_time_loc(file_data, 'meta_info_loc', 'new')
        logging.info('正在修改文件[每个大包时间]...')
        idx = 1
        for item in pack_data_time_loc_list:
            loc_s = int(item[0])
            loc_e = loc_s + int(item[1])
            # 缓存本包结束时间:为与开始时间之差的毫秒数
            pack_end_time_diff = int.from_bytes(file_data[loc_e:loc_e+4], byteorder='big')
            if idx == 1:
                b_time_array, timestamp = b_beginTime_array, new_begin_timestamp
            else:
                b_time_array = inttimestampTobytes(timestamp)
            new_file_data = update_file_data(new_file_data, b_time_array, loc_s, loc_e)
            timestamp += pack_end_time_diff
            idx += 1
        if file_type == 'stp':   # 则还需要修改计步文件中每个点的时间
            metaInfo_length = hexToInt(bytesToHexString(file_data[78:80]))
            meta_info_list, \
            metaInfo_item_list, \
            metaExtInfo_list = parse_MetaInfo(file_data[78:metaInfo_length + 80])
            data_length = hexToInt(bytesToHexString(file_data[metaInfo_length + 80:metaInfo_length + 84]))
            pack_data = new_file_data[metaInfo_length + 84:metaInfo_length + 84 + data_length]
            new_pack_data = only_parse_Stp_PackData(pack_data, len(pack_data), metaInfo_item_list,
                                                    old_begin_timestamp, new_begin_timestamp)
            new_file_data = new_file_data[0:metaInfo_length + 84]+new_pack_data
        logging.info('修改文件[大包数据时间]结束...')
        return new_file_data, file_end_time
    except Exception as e:
        logging.error(f'修改文件时间数据出错啦！原因为：{e}')
        return None, None

# 解析计步文件中的包数据
def only_parse_Stp_PackData(pack_data, data_len, item_data_list, old_begin_timestamp, new_begin_timestamp):
    try:
        index = 0
        i = 1
        logging.info(f'开始解析文件包体数据...')
        while data_len > 0:
            # 解析包头信息:20bytes
            packData_len = hexToInt(bytesToHexString(pack_data[index+12:index+16]))
            index += 20
            data_len -= 20
            # 解析包体数据
            packData_len2 = packData_len
            while packData_len > 0:
                for item in item_data_list:
                    if item['dataLen'] > 0:
                        value = pack_data[index:index+item['dataLen']]
                        stmp_list = parse_SmallPackData(item['type'], item['version'], value)
                        if item['type'] == '13':
                            stmp = int(stmp_list[0]) - old_begin_timestamp
                            b_time_array = inttimestampTobytes(new_begin_timestamp + abs(stmp))
                            pack_data = update_file_data(pack_data, b_time_array, index, index+item['dataLen'])
                        index += item['dataLen']
                        packData_len -= item['dataLen']
            data_len -= packData_len2
            i += 1
        return pack_data
    except Exception as e:
        return None

def get_file_type(item):
    base_name = os.path.basename(item)
    file_type = ''
    if str(base_name).__contains__('.'):
        new_item = str(base_name).split('.')
        file_type = new_item[len(new_item) - 1]
        if new_item[len(new_item) - 1].__contains__('_old'):
            file_type = str(new_item[len(new_item) - 1]).split('_')[0]
    return file_type

def get_new_time_data(b_time, e_time, modify_data, modify_ty):
    b_time_stamp = int.from_bytes(b_time, byteorder='big')
    e_time_stamp = int.from_bytes(e_time, byteorder='big')
    time_diff = e_time_stamp - b_time_stamp
    new_data = get_timedate('', modify_data, modify_ty)
    timestamp = get_timestamp(new_data) + time_diff
    b_time_array = inttimestampTobytes(timestamp)
    return b_time_array

def get_time_data(raw_data, modify_data, modify_ty):
    new_data = get_timedate(raw_data, modify_data, modify_ty)
    timestamp = get_timestamp(new_data)
    timestamp += (int.from_bytes(raw_data, byteorder='big') % 1000)
    b_time_array = inttimestampTobytes(timestamp)
    return b_time_array, timestamp

def timestampToDate(timestamp):
    timeStamp = float(timestamp/1000)
    timeArray = time.localtime(timeStamp)
    otherStyleTime = time.strftime('%Y%m%d%H%M%S', timeArray)
    return otherStyleTime

def get_timedate(raw_time, modify_data, modify_ty):
    if modify_ty == 'all':
        datetime_type = datetime.datetime.fromtimestamp(hexToInt(bytesToHexString(raw_time)) // 1000)
        beijin_time = datetime_type.strftime("%Y-%m-%d %H:%M:%S")
        stmp = str(beijin_time).split(' ')
        modify_date = modify_data[0:4] + '-' + modify_data[4:6] + '-' + modify_data[6:8]
        new_date = modify_date + ' ' + stmp[1]
    else:
        dt_obj = datetime.datetime.strptime(modify_data, '%Y%m%d%H%M%S')
        new_date = dt_obj.strftime('%Y-%m-%d %H:%M:%S')
    return new_date

def get_timestamp(new_time):
    # 将字符串时间转换为13位时间戳之方法一
    struct_time = time.strptime(new_time, '%Y-%m-%d %H:%M:%S')
    timestamp = int(time.mktime(struct_time))*1000
    # 将字符串时间转换为13位时间戳之方法二
    # timeArray = datetime.datetime.strptime(new_time, '%Y-%m-%d %H:%M:%S')
    # strtime = timeArray.strftime('%Y-%m-%d %H:%M:%S.%f')
    # timeArray = datetime.datetime.strptime(strtime, '%Y-%m-%d %H:%M:%S.%f')
    # timestamp = int(time.mktime(timeArray.timetuple())*1000.0+timeArray.microsecond / 1000.0)
    return timestamp

def get_packData_time_loc(file_data, item, p_ty='new'):
    pack_time_list = []
    mateInfo_loc_s, mateInfo_loc_e  = get_data_loc(item, p_ty)
    mateInfo_length = hexToInt(bytesToHexString(file_data[mateInfo_loc_s:mateInfo_loc_e]))
    data_length = hexToInt(bytesToHexString(file_data[mateInfo_length + 80:mateInfo_length + 84]))
    s_loc = mateInfo_length+84
    while data_length > 0:
        pack_time_list.append([s_loc, 8])
        packData_len = hexToInt(bytesToHexString(file_data[s_loc+16:s_loc+20]))
        s_loc += (packData_len + 20)
        data_length -= (packData_len + 20)
    return pack_time_list

def make_new_file_data(file_all_data, pack_info, type='old', mate_len=0):
    pack_data = bytearray()
    file_end = bytearray()
    for i in range(len(pack_info)):
        pack_data += pack_info[i]['pack_head']
        pack_data += pack_info[i]['pack_data']
        if type == 'old':
            if i == 0:
                pack_len = 151
            else:
                pack_len += len(pack_info[i - 1]['pack_head']) + len(pack_info[i - 1]['pack_data'])
            file_end += (pack_len).to_bytes(4, byteorder='big')
    # 修改加密标志为0,未加密
    enc_flg = (0).to_bytes(1, byteorder='little')
    if type == 'old':
        # 修改文件尾开始位置，4个字节
        file_end_loc = (151 + len(pack_data)).to_bytes(4, byteorder='big')
        file_head = file_all_data[0:107] + enc_flg + file_all_data[108:110] + file_end_loc + file_all_data[114:151]
        new_file = file_head + pack_data + file_end
    elif type == 'new':
        data_len = len(pack_data).to_bytes(4, byteorder='big')
        # print(f'data_len={data_len}')
        file_head = file_all_data[0:74] + enc_flg + file_all_data[75:mate_len + 80] + data_len
        new_file = file_head + pack_data
    return new_file

def make_modify_data(obj, file_name_list, user_id, device_number, file_time, modify_ty, dll_obj):
    try:
        for item in file_name_list:
            file_data = get_file_data(item)
            file_type = ''
            if (str(item).__contains__('stp')):
                file_type = 'stp'
                encrypt_flg, ty = validate_fileIsEncrypt(file_data)
                if encrypt_flg != 0:  # 已加密
                    if ty == 'new':
                        logging.info(f'文件{enc_file_name}已加密，正在解密文件...')
                        pack_data_list, mateInfo_length = dll_obj.decrypt_new_protocol_file(file_data, ty)
                        logging.info(f'文件{enc_file_name}解密结束...')
                        # 保存解密后的数据
                        file_data = make_new_file_data(file_data, pack_data_list, item, mateInfo_length)
            logging.info('开始修改数据...')
            if file_data is not None:
                if int(file_data[0]) < 10:  # 旧协议
                    show_message(obj, '消息提示', f'文件:{item}为旧文件协议，本次将跳过该文件的修改...', 'xxx')
                else:
                    ty = 'new'
                    modify_file_time_flg = False   # 修改文件时间标记，默认为False
                    new_data = file_data
                    file_end_time = ''
                    if len(user_id) != 0:  # 修改USERID数据
                        logging.info(f'正在修改USEID数据为：{user_id}')
                        mod_data = str.encode(user_id, 'utf-8')  # 需要将字符串数据转换为bytearray
                        loc_s, loc_e = get_data_loc('user_id_loc', ty)
                        new_data = update_file_data(new_data, mod_data, loc_s, loc_e)
                    if len(device_number) != 0:  # 修改设备编号数据
                        logging.info(f'正在修改设备编号数据为：{device_number}')
                        mod_data = str.encode(device_number, 'utf-8')  # 需要将字符串数据转换为bytearray
                        loc_s, loc_e = get_data_loc('device_num_loc', ty)
                        new_data = update_file_data(new_data, mod_data, loc_s, loc_e)
                    if len(file_time) != 0:  # 修改文件时间数据
                        logging.info(f'正在修改文件时间数据为：{file_time}')
                        new_data, file_end_time = modify_file_time(new_data, file_time, ty, modify_ty, file_type)
                        modify_file_time_flg = True
                    logging.info('修改数据结束...')
                    # 先修改原文件名,在后面加”_old“，前提是原文件名中不包含”_old"后缀
                    if not str(item).__contains__('_old'):
                        if not os.path.exists(item+'_old'):
                            os.rename(item, item + '_old')
                    else:   # 如果包含，则新保存的文件名则要支掉“_old”后缀
                        item = str(item).replace('_old', '')
                    # 将修改后的数据，保留原文件名不变，然后再写入文件中去
                    if not new_data is None:
                        save_modifyData_To_file(obj, new_data, item, '', modify_ty, modify_file_time_flg, file_time, file_end_time)
    except Exception as e:
        show_message(obj, '消息提示',f'出错啦，原因为：{e}...', 'err')

# 将数据写入到文件
def save_modifyData_To_file(obj, file_data, old_path, prefix, ty, modify_file_time_flg, new_time, file_end_time):
    try:
        (file_path, file_name) = os.path.split(old_path)
        new_file_name = file_name
        print(f'new_file_name1={new_file_name}')
        if modify_file_time_flg:   # 表示存在修改时间，则按新时间进行文件重命名
            # 包含"_"有可能是jksense设备采集的数据与手表计步文件的数据
            # jksense设备采集的数据，需要重命名时：同时修改开始时间与结束时间
            # 手表计步文件的数据，需要重命名时：仅修改开始时间
            if str(file_name).__contains__('_'):
                file_name_list = file_name.split('_')
                if ty == 'one':
                    new_file_name = f'{new_time}'
                    if 'jks' in file_name_list[len(file_name_list)-1]:
                        new_file_name += f'_{file_end_time}_{file_name_list[len(file_name_list)-1]}'
                    elif 'stp' in file_name_list[len(file_name_list)-1]:   # 表示计步文件,仅修改开始时间
                        new_file_name += f'_{file_name_list[len(file_name_list)-1]}'
                else:
                    new_file_name = f'{new_time[0:8] + file_name_list[0][8:]}'
                    if 'jks' in file_name_list[len(file_name_list) - 1]:
                        new_file_name += f'_{file_end_time}_{file_name_list[len(file_name_list)-1]}'
                    elif 'stp' in file_name_list[len(file_name_list)-1]:   # 表示计步文件,仅修改开始时间
                        new_file_name += f'_{file_name_list[len(file_name_list)-1]}'
            else:
                if ty=='one':
                    new_file_name = new_time + file_name[14:]
                else:
                    new_file_name = new_time[0:8]+file_name[8:]
            print(f'new_file_name2={new_file_name}')
        if len(prefix) == 0:
            new_path_and_file = file_path + '/' + new_file_name
        else:
            new_path_and_file = file_path + '/' + prefix + '_' + new_file_name
        with open(new_path_and_file, 'wb') as f:
            f.write(file_data)
            f.close()
            if ty == 'one':
                msg = f'修改文件数据成功，文件已按重命名规则进行重命名并保存完成...'
                show_message(obj, '消息提示', msg, 'info')
            return True
    except Exception as e:
        err = f'修改文件数据失败！原因为：{e}...'
        show_message(obj, '消息提示', err, 'err')
        return False

def disope_table_head(v):
    stmp = v
    s = []
    loc_list = []
    l_char = '{'
    r_char = '}'
    if str(stmp).__contains__(l_char) and str(stmp).__contains__(r_char):
        c = v.count(l_char)
        for i in range(c):
            loc1 = v.find(l_char)
            loc2 = v.find(r_char)
            v2 = v[loc1 + 1: loc2]
            v3 = str(v2).split(' ')
            s.append(v3.copy())
            v3.clear()
            v = v[loc2 + 1:]
        ss = str(stmp).split(' ')
        j = 1
        k = 0
        for item in ss:
            if str(item).__contains__(l_char) and str(item).__contains__(r_char):
                loc_list.append(j)
            else:
                if str(item).__contains__(l_char):
                    loc_list.append(j - k)
                if str(item).__contains__(r_char):
                    k += 1
            j += 1
    else:
        s.append(str(v).split(' '))
    return s, loc_list


# 匹配{}之间的数据
def extract_data_between_braces(text):
    # 正则表达式匹配大括号内的内容
    pattern = r"\{([^}]+)\}"
    matches = re.findall(pattern, text)
    return matches

# 验证日期是否合法
def is_valid_date(date_text):
    try:
        new_date = date_text[0:4]+'-'+date_text[4:6]+'-'+date_text[6:]
        datetime.datetime.strptime(new_date, '%Y-%m-%d')  # 日期格式
    except ValueError:
        return False
    return True

# 验证时间是否合法
def is_valid_time(date_text):
    try:
        new_time = date_text[0:2] + ':' + date_text[2:4] + ':' + date_text[4:]
        datetime.datetime.strptime(new_time, '%H:%M:%S')  # 时间格式
    except ValueError:
        return False
    return True

def disp_file_path(obj):
    file_path = obj.get()
    file_path = file_path.strip('{}')  # 去年首尾的{},如果不存在也不会报错
    return file_path

def get_all_file_types():
    logging.info('正在获取支持的文件类型...')
    all_file_types = ''
    for k, v in file_types_dic.items():
        all_file_types += v
    return all_file_types

# 将13时间戳转换为bytes
def inttimestampTobytes(v):
    return v.to_bytes(length=8, byteorder='big', signed=False)

def bytesToString(bs):
    '''
        bytes to string
        eg:
        b'0123456789ABCDEF0123456789ABCDEF'
        '0123456789ABCDEF0123456789ABCDEF'
    '''
    return bytes.decode(bs,encoding='utf8')

def stringTobytes(str):
    '''
        string to bytes
        eg:
        '0123456789ABCDEF0123456789ABCDEF'
        b'0123456789ABCDEF0123456789ABCDEF'
    '''
    return bytes(str,encoding='utf8')

def bytesToShort(byte_data):
    s_list = []
    step = 2
    for i in range(0, len(byte_data), step):
        v1 = byte_data[i:i + step]
        # v2 = struct.unpack('<h', v1)[0]    # 'h':2字节整数，其中'>'为大端，'<'为小端
        v2 = struct.unpack('@e', v1)[0]      # 'e':2字节浮点数，其中'>'为大端，'<'为小端
        if math.isnan(v2):  # 如果是nan，则赋为0
            v2 = 0.0
        v3 = "{:.5f}".format(v2)  # 四舍五入保留了5位小数不足位补0
        s_list.append(v3)
    return s_list

# ecg数据，转换为2字节整数
def bytes_To_TowInt(byte_data):
    s_list = []
    step = 2
    for i in range(0, len(byte_data), step):
        v1 = byte_data[i:i+step]
        v2 = struct.unpack('>h', v1)[0]   # 'h'2字节整数, 需要使用">"大端
        # v3 = "{:.5f}".format(v2)  # 四舍五入保留了5位小数不足位补0
        s_list.append(v2)
    return s_list

# 四字节浮点数
def bytesToFloat(byte_data):
    ba = bytearray()
    k = 0
    s_list = []
    for i in range(len(byte_data)):
        ba.append(byte_data[i])
        k += 1
        if k % 4 == 0:
            v2 = struct.unpack('!f',ba)[0]
            v3 = "{:.5f}".format(v2)  # 四舍五入保留了5位小数不足位补0
            s_list.append(v3)
            ba.clear()
    return s_list


def hexStringTobytes(str):
    '''
        hex string to bytes
        eg:
        '01 23 45 67 89 AB CD EF 01 23 45 67 89 AB CD EF'
        b'\x01#Eg\x89\xab\xcd\xef\x01#Eg\x89\xab\xcd\xef'
    '''
    str = str.replace(" ", "")
    return bytes.fromhex(str)


def bytesToHexString(bs):
    '''
        bytes to hex string
        eg:
        b'\x01#Eg\x89\xab\xcd\xef\x01#Eg\x89\xab\xcd\xef'
        '01 23 45 67 89 AB CD EF 01 23 45 67 89 AB CD EF'
    '''
    # hex_str = ''
    # for item in bs:
    #     hex_str += str(hex(item))[2:].zfill(2).upper() + " "
    # return hex_str
    return ''.join(['%02X' % b for b in bs])

def bytes_to_float(bytes, byteorder='big'):
    return struct.unpack(f'{byteorder}f', bytes)[0]

def hexToInt_list(v, n=4):
    s_list = []
    for i in range(0, len(v), n):
        vv = int(v[i:i+n], 16)
        s_list.append(vv)
    return s_list

def hexToInt(v):
    '''
        hex to int
        eg:
        '0a'
        10
    '''
    try:
        return int(v, 16)
    except Exception as e:
        print(e)

def config_logging(log_path, file_name, console_level: int=logging.INFO, file_level: int=logging.DEBUG):
    '''
    管理日志数据
    :param log_file_name:
    :param console_level:
    :param file_level:
    :return:
    '''
    '''
    Value	      Type of interval
        S	   等待1秒切换到一个新日志文件
        M	   等待1分钟切换到一个新日志文件
        H	   等待1小时切换到一个新日志文件
        D	   等待1天切换到一个新日志文件
        W	   等待1星期切换到一个新日志文件 (0=Monday)
    midnight   每天0点切换到新的日志
    '''
    formatter = logging.Formatter(u'[%(asctime)s - %(levelname)s - %(funcName)s] --> %(message)s')
    # 创建文件对象
    # file_handler = logging.FileHandler(log_file_name, mode='a', encoding="utf8")
    # 修改：采用每天0时自动换日志文件的方式进行创建文件对象 2023-12-14
    if not os.path.exists(log_path):  # 不存在文件夹，则先创建
        os.mkdir(log_path)
    file_handler = TimedRotatingFileHandler(log_path+file_name, when="midnight", encoding="utf8")
    file_handler.setFormatter(formatter)
    file_handler.setLevel(file_level)
    # 创建控制台对象
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    console_handler.setLevel(console_level)
    logging.basicConfig(
        level=min(console_level, file_level),
        handlers=[file_handler, console_handler])

def csv_file_write(csv_file_name, write_row_data):
    """ csv创建并写入数据
    写文件模式
        w模式，文件不存在，会创建新文件；反之存在，会清空原有内容
        a模式，文件不存在，会创建新文件；反之存在，会在原有内容后面继续追加写入
    """
    try:
        # encoding='utf-8-sig'为解决中文乱码
        with open(csv_file_name, 'a', encoding='utf-8-sig', newline='') as f:
            csv_write = csv.writer(f)
            csv_write.writerow(write_row_data)
            f.close()
    except Exception as e:
        logging.error("打开文件["+csv_file_name+"]异常..." + str(e))
        return None

def comp_file(file1, file2):
    '''
    文件比较
    :param file1:
    :param file2:
    :return:
    '''
    l_file_data = get_file_data(file1)
    r_file_data = get_file_data(file2)
    # 组合写入csv的每条记录数据
    content_item = []
    content_item.append(file1)
    content_item.append(file2)
    l_file_data_array = list(l_file_data)
    r_file_data_array = list(r_file_data)
    if l_file_data_array == r_file_data_array:
        content_item.append('OK')
        content_item.append('无差异')
    else:
        # comp_result = DeepDiff(l_file_data_array, r_file_data_array, significant_digits=1)
        content_item.append('NO')
        content_item.append('存在差异')
    csv_file_write("_comp_result.csv", content_item)

def get_file_list(path):
    '''
    获取指定文件目录下的所有文件，包括目录及文件
    :param path:
    :return:List
    '''
    logging.info('读取目录' + path + '中的文件数据...')
    file_list = []
    for i, j, k in os.walk(path):
        file_list.append([i, j, k])
    # logging.info('读取的文件列表数据' + str(file_list) + '...')
    return file_list

def save_test_result_to_excel(src_file_name, test_result_list, dst_file_name):
    '''
    保存测试结果到EXCEL
    :param src_file_name:
    :param test_result_list:
    :param dst_file_name:
    :return:
    '''
    try:
        el_data = openpyxl.load_workbook(src_file_name)
        option_table = el_data.active
        list_data = list(option_table.iter_rows(values_only=True))
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Sheet1'
        # 首先写入表头
        worksheet.append(list_data[0])
        for item in test_result_list:
            worksheet.append(tuple(item))
        workbook.save(dst_file_name)
        logging.info('写测试结果到Excel文件成功...')
    except Exception as e:
        logging.error('写测试结果到文件出错啦...' + str(e))
    # el_data.close()
    # workbook.close()